#WITHOUT PARALLELISM
import time
startTime = time.time()
from classstatisticsfinal2 import Variance, Correlation, Regression
import openpyxl
path="newsheet.xlsx"
wbobj=openpyxl.load_workbook(path)
sheet =wbobj.active
data=[]
for row in data:
    sheet.append(row)
row=sheet[2]
rowlist=[row[x].value for x in range(2,len(row))]
print("original list is: \n")
print(rowlist)
print("\n")
def quicksort(row_list):
    if row_list==[]:
        return []
    pivot=row_list[0]
    less=[]
    greater=[]
    for item in row_list[1:]:
        if item<pivot:
            less.append(item)
        else:
            greater.append(item)
    sortedlist=quicksort(less)+[pivot]+quicksort(greater)
    return sortedlist
finallist=quicksort(rowlist)
print("sorted list is: \n")
print(finallist)
print("\n")

column = sheet['B']
columnlist=[column[x].value for x in range(1,len(column))]
print("column 1 is: \n")
print(columnlist)
print("\n")
column = sheet['C']
columnlist1=[column[x].value for x in range(1,len(column))]
print("column 2 is: \n")
print(columnlist1)
print("\n")

mean1=sum(columnlist)/len(columnlist)
mean2=sum(columnlist1)/len(columnlist1)

##############################################################

#VARIANCE
a = Variance(finallist, rowlist, mean1, mean2)
#POPULATION VARIANCE
print("Variance sample: ")
print(a.variancesample(rowlist))
#SAMPLE VARIANCE
print("Variance sample higher order: ")
print(a.higerorder_variancesample(rowlist))
#POPULATION VARIANCE HIGHER ORDER RECURSION
print("Variance population: ")
print(a.variance(rowlist))
#SAMPLE VARIANCE HIGHER ORDER RECURSION
print("Variance population higher order: ")
print(a.higerorder_variance(rowlist))

#CORRELATION
d = Correlation(columnlist, columnlist1, mean1, mean2)
#PEARSON CORRELATION
print("Pearson Correlation: ")
print(d.correlation(columnlist, columnlist1))
#SAMPLE CORRELATION 
print("Sample Correlation: ")
print(d.samplecorrelation(columnlist, columnlist1))
#LINEAR CORRELATION
print("Linear Correlation: ")
print(d.lincorrelation(columnlist, columnlist1))
#POPULATION CORRELATION
print("Population Correlation: ")
print(d.popcorrelation(columnlist, columnlist1))

#REGRESSION
g = Regression(columnlist, columnlist1, mean1, mean2)
print("Regression: ")
print(g.regression(columnlist, columnlist1))



#PRINT TIME
executionTime = (time.time() - startTime)
print('Execution time in seconds for without parallelism: ' + str(executionTime))


###################################################################
# WRITE TO SHEET

#import the writer
import xlwt
#import the reader
import xlrd
#open the sussex results spreadsheet
#book = xlrd.open_workbook("newsheet.xlsx")
path="newsheet.xlsx"
wbobj=openpyxl.load_workbook(path)
sheet =wbobj.active
# WRITE POPULATION VARIANCE 

cell1 = sheet.cell(row = 13, column = 1)
cell1.value = "Variance sample:"
cell2 = sheet.cell(row= 13 , column = 2)
cell2.value = a.variancesample(rowlist)
# WRITE SAMPLE VARIANCE


cell1 = sheet.cell(row = 14, column = 1)
cell1.value = "Variance sample higher order:"
cell2 = sheet.cell(row= 14 , column = 2)
cell2.value = a.higerorder_variancesample(rowlist)


# WRITE POPULATION VARIANCE HIGHER ORDER RECURSION

cell1 = sheet.cell(row = 15, column = 1)
cell1.value = "Variance population:"
cell2 = sheet.cell(row= 15 , column = 2)
cell2.value = a.variance(rowlist)

# WRITE SAMPLE VARIANCE HIGHER ORDER RECURSION


cell1 = sheet.cell(row = 16, column = 1)
cell1.value = "Variance population higher order:"
cell2 = sheet.cell(row= 16, column = 2)
cell2.value = a.higerorder_variance(rowlist)

# WRITE PEARSON CORRELATION


cell1 = sheet.cell(row = 17, column = 1)
cell1.value = "Pearson Correlation::"
cell2 = sheet.cell(row= 17, column = 2)
cell2.value = d.correlation(columnlist, columnlist1)

# WRITE SAMPLE CORRELATION 


cell1 = sheet.cell(row = 18, column = 1)
cell1.value = "Sample Correlation:"
cell2 = sheet.cell(row= 18, column = 2)
cell2.value = d.samplecorrelation(columnlist, columnlist1)

# WRITE LINEAR CORRELATION


cell1 = sheet.cell(row = 19, column = 1)
cell1.value = "Linear Correlation:"
cell2 = sheet.cell(row= 19, column = 2)
cell2.value = d.lincorrelation(columnlist, columnlist1)

# WRITE POPULATION CORRELATION


cell1 = sheet.cell(row = 20, column = 1)
cell1.value = "Population Correlation:"
cell2 = sheet.cell(row= 20, column = 2)
cell2.value = d.popcorrelation(columnlist, columnlist1)

# WRITE REGRESSION

cell1 = sheet.cell(row = 21, column = 1)
cell1.value = "Regression: "
cell2 = sheet.cell(row= 21, column = 2)
cell2.value = g.regression(columnlist, columnlist1)


cell1 = sheet.cell(row = 22, column = 1)
cell1.value = "Time: "
cell2 = sheet.cell(row= 22, column = 2)
cell2.value = str(executionTime)

wbobj.save("newsheetnewnew1.xlsx")








