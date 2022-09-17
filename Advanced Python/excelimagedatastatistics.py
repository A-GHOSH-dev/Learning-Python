
import pandas as pd
#import io,os
import numpy as np
import openpyxl
import matplotlib.pyplot as plt

#convert csv to excel 
# Reading the csv file
df_new = pd.read_csv('sheetf1.csv')
 
# saving xlsx file
ANU = pd.ExcelWriter('ananyaimagedata.xlsx')
df_new.to_excel(ANU, index=False)
 
ANU.save()


from classstatisticsfinal2 import Variance, Correlation, Regression

path="ananyaimagedata.xlsx"
wbobj=openpyxl.load_workbook(path)
sheet =wbobj.active
data=[]
for row in data:
    sheet.append(row)
row=sheet[2]
rowlist=[row[x].value for x in range(2,len(row))]
print("original list is: \n")
print(rowlist)
print("\n")
def quicksort(row_list):
    if row_list==[]:
        return []
    pivot=row_list[0]
    less=[]
    greater=[]
    for item in row_list[1:]:
        if item<pivot:
            less.append(item)
        else:
            greater.append(item)
    sortedlist=quicksort(less)+[pivot]+quicksort(greater)
    return sortedlist
finallist=quicksort(rowlist)
print("sorted list is: \n")
print(finallist)
print("\n")
column = sheet['C']
columnlist=[column[x].value for x in range(1,len(column))]
print("column 1 is: \n")
print(columnlist)
print("\n")
column = sheet['D']
columnlist1=[column[x].value for x in range(1,len(column))]
print("column 2 is: \n")
print(columnlist1)
print("\n")

mean1=sum(columnlist)/len(columnlist)
mean2=sum(columnlist1)/len(columnlist1)
##############################################################
#VARIANCE
a = Variance(finallist, rowlist, mean1, mean2)
#POPULATION VARIANCE
print("Variance sample: ")
print(a.variancesample(rowlist))
#SAMPLE VARIANCE
print("Variance sample higher order: ")
print(a.higerorder_variancesample(rowlist))
#POPULATION VARIANCE HIGHER ORDER RECURSION
print("Variance population: ")
print(a.variance(rowlist))
#SAMPLE VARIANCE HIGHER ORDER RECURSION
print("Variance population higher order: ")
print(a.higerorder_variance(rowlist))
#CORRELATION
d = Correlation(columnlist, columnlist1, mean1, mean2)
#PEARSON CORRELATION
print("Pearson Correlation: ")
print(d.correlation(columnlist, columnlist1))
#SAMPLE CORRELATION 
print("Sample Correlation: ")
print(d.samplecorrelation(columnlist, columnlist1))
#LINEAR CORRELATION
print("Linear Correlation: ")
print(d.lincorrelation(columnlist, columnlist1))
#POPULATION CORRELATION
print("Population Correlation: ")
print(d.popcorrelation(columnlist, columnlist1))
#REGRESSION
g = Regression(columnlist, columnlist1, mean1, mean2)
print("Regression: ")
print(g.regression(columnlist, columnlist1))
###################################################################
# WRITE TO SHEET
path="ananyawritesheet.xlsx"
wbobj=openpyxl.load_workbook(path)
sheet =wbobj.active
# WRITE POPULATION VARIANCE 
cell1 = sheet.cell(row = 13, column = 1)
cell1.value = "Variance sample:"
cell2 = sheet.cell(row= 13 , column = 2)
cell2.value = a.variancesample(rowlist)
# WRITE SAMPLE VARIANCE
cell1 = sheet.cell(row = 14, column = 1)
cell1.value = "Variance sample higher order:"
cell2 = sheet.cell(row= 14 , column = 2)
cell2.value = a.higerorder_variancesample(rowlist)
# WRITE POPULATION VARIANCE HIGHER ORDER RECURSION
cell1 = sheet.cell(row = 15, column = 1)
cell1.value = "Variance population:"
cell2 = sheet.cell(row= 15 , column = 2)
cell2.value = a.variance(rowlist)
# WRITE SAMPLE VARIANCE HIGHER ORDER RECURSION
cell1 = sheet.cell(row = 16, column = 1)
cell1.value = "Variance population higher order:"
cell2 = sheet.cell(row= 16, column = 2)
cell2.value = a.higerorder_variance(rowlist)
# WRITE PEARSON CORRELATION
cell1 = sheet.cell(row = 17, column = 1)
cell1.value = "Pearson Correlation::"
cell2 = sheet.cell(row= 17, column = 2)
cell2.value = d.correlation(columnlist, columnlist1)
# WRITE SAMPLE CORRELATION 
cell1 = sheet.cell(row = 18, column = 1)
cell1.value = "Sample Correlation:"
cell2 = sheet.cell(row= 18, column = 2)
cell2.value = d.samplecorrelation(columnlist, columnlist1)
# WRITE LINEAR CORRELATION
cell1 = sheet.cell(row = 19, column = 1)
cell1.value = "Linear Correlation:"
cell2 = sheet.cell(row= 19, column = 2)
cell2.value = d.lincorrelation(columnlist, columnlist1)
# WRITE POPULATION CORRELATION
cell1 = sheet.cell(row = 20, column = 1)
cell1.value = "Population Correlation:"
cell2 = sheet.cell(row= 20, column = 2)
cell2.value = d.popcorrelation(columnlist, columnlist1)
# WRITE REGRESSION
cell1 = sheet.cell(row = 21, column = 1)
cell1.value = "Regression: "
cell2 = sheet.cell(row= 21, column = 2)
cell2.value = g.regression(columnlist, columnlist1)
#SAVE NEW FILE
wbobj.save("ananyaimagenew.xlsx")

print('\n')
print('\n')

#NUMPY AND PANDAS

# csv to dataframe
print('csv to dataframe')
dataframe1 = pd.read_csv('sheetf1.csv')
print('\n')
print(dataframe1)
#find variance of all columns of the dataframe using numpy and pandas in 3 ways
print('find variance of all columns of the dataframe using numpy and pandas in 3 ways')
print('\n')
var1=dataframe1[['Area','equivalent_diameter','orientation','MajorAxisLength','MinorAxisLength','Perimeter','MinIntensity','MeanIntensity','MaxIntensity']].var()
print(dataframe1[['Area','equivalent_diameter','orientation','MajorAxisLength','MinorAxisLength','Perimeter','MinIntensity','MeanIntensity','MaxIntensity']].var())
print('\n')
var2= np.var(dataframe1[['Area','equivalent_diameter','orientation','MajorAxisLength','MinorAxisLength','Perimeter','MinIntensity','MeanIntensity','MaxIntensity']])
print(np.var(dataframe1[['Area','equivalent_diameter','orientation','MajorAxisLength','MinorAxisLength','Perimeter','MinIntensity','MeanIntensity','MaxIntensity']]))
print('\n')
var3= np.var(dataframe1[['Area','equivalent_diameter','orientation','MajorAxisLength','MinorAxisLength','Perimeter','MinIntensity','MeanIntensity','MaxIntensity']],ddof=1)
print(np.var(dataframe1[['Area','equivalent_diameter','orientation','MajorAxisLength','MinorAxisLength','Perimeter','MinIntensity','MeanIntensity','MaxIntensity']],ddof=1))
print('\n')

#find Correlation of all columns of the dataframe using numpy and pandas in 3 ways
print('find Correlation of all columns of the dataframe using numpy and pandas in 3 ways')
print('\n')
corrrelation1    = dataframe1.corr(method="pearson");
print("Pearson correlation coefficient:");
print(corrrelation1); 
print('\n')
corrrelation2    = dataframe1.corr(method="kendall");
print("Kendall Tau correlation coefficient:");
print(corrrelation2); 
print('\n')
corrrelation3   = dataframe1.corr(method="spearman");
print("Spearman rank correlation:");
print(corrrelation3);

#find regression of all columns of the dataframe using numpy and pandas in 3 ways

path="ananyawritesheet1.xlsx"
wbobj=openpyxl.load_workbook(path)
sheet =wbobj.active
# Variance 1
cell1 = sheet.cell(row = 13, column = 1)
cell1.value = "Variance:"
cell2 = sheet.cell(row= 13 , column = 2)
cell2.value = str(var1)
# Variance 2
cell1 = sheet.cell(row = 14, column = 1)
cell1.value = "Variance:"
cell2 = sheet.cell(row= 14 , column = 2)
cell2.value = str(var2)
# Variance 3
cell1 = sheet.cell(row = 15, column = 1)
cell1.value = "Variance:"
cell2 = sheet.cell(row= 15 , column = 2)
cell2.value = str(var3)
# correlation 1
cell1 = sheet.cell(row = 16, column = 1)
cell1.value = "Correlation1:"
cell2 = sheet.cell(row= 16 , column = 2)
cell2.value = str(corrrelation1)
# correlation 2
cell1 = sheet.cell(row = 17, column = 1)
cell1.value = "Correlation2:"
cell2 = sheet.cell(row= 17 , column = 2)
cell2.value = str(corrrelation2)
# correlation 3
cell1 = sheet.cell(row = 18, column = 1)
cell1.value = "Correlation3:"
cell2 = sheet.cell(row= 18 , column = 2)
cell2.value = str(corrrelation3)


wbobj.save("ananyaimagenewnew.xlsx")

#MATPLOTLIB

plt.plot(var1)
plt.show()

plt.plot(var2)
plt.show()

plt.plot(var3)
plt.show()

plt.plot(corrrelation1)
plt.show()

plt.plot(corrrelation2)
plt.show()

plt.plot(corrrelation3)
plt.show()